from django.http import HttpResponse
from django.shortcuts import redirect
from django.views import View
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


class DownloadExcelTemplateView(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return redirect("accounts:login")

        wb = Workbook()
        ws = wb.active
        ws.title = "Product Template"

        headers = [
            "Name",
            "Category",
            "Brand",
            "SKU",
            "Serial Number",
            "Price",
            "Description",
            "Branch (Code)",
            "Local SKU",
            "Rack Number",
            "Shelf Number",
            "Datasheet Filename",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font, cell.fill, cell.alignment = (
                Font(bold=True, color="FFFFFF"),
                PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid"
                ),
                Alignment(horizontal="center"),
            )

        sample_data = [
            "Sample Product",
            "Electronics",
            "Sample Brand",
            "SKU001",
            "SN123456",
            "99.99",
            "Sample product description",
            "IIA",
            "L-SKU001",
            "A1",
            "B2",
            "widget2000.pdf",
        ]
        for col, value in enumerate(sample_data, 1):
            cell = ws.cell(row=2, column=col, value=value)
            cell.font = Font(italic=True, color="666666")

        for column in ws.columns:
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column[0].column_letter].width = min(
                max_length + 2, 50
            )

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            'attachment; filename="product_upload_template.xlsx"'
        )
        wb.save(response)
        return response


def download_excel_template(request):
    return DownloadExcelTemplateView().get(request)
